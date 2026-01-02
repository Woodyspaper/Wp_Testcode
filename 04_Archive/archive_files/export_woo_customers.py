"""
Export WooCommerce customers to CSV (default) or Excel

Usage:
    python export_woo_customers.py          # CSV (default)
    python export_woo_customers.py csv      # CSV
    python export_woo_customers.py excel    # Excel with formatting
    python export_woo_customers.py both     # Both formats
"""
import os
import csv
import sys
from datetime import datetime

# Add project root to Python path
project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if project_root not in sys.path:
    sys.path.insert(0, project_root)

from woo_client import WooClient


def fetch_all_customers():
    """
    Fetch all customers from WooCommerce.
    
    IMPORTANT: WooCommerce has a known bug where the /customers endpoint
    doesn't return all customers. To work around this, we:
    1. Fetch from /customers with role=all
    2. Also extract unique customers from /orders
    3. Merge and deduplicate by ID
    """
    print("=" * 60)
    print("FETCHING WOOCOMMERCE CUSTOMERS")
    print("=" * 60)
    
    woo = WooClient()
    customers_by_id = {}
    
    # Step 1: Fetch from customer list (with role=all to get all roles)
    print("\n[1/2] Fetching from customer list API...")
    page = 1
    per_page = 100
    
    while True:
        url = woo._url("/customers")
        # Use role=all to get customers with any role (not just 'customer')
        response = woo.session.get(url, params={"page": page, "per_page": per_page, "role": "all"}, timeout=60)
        if not response.ok:
            print(f"  Error fetching page {page}: {response.status_code}")
            break
        customers = response.json()
        if not customers:
            break
        for c in customers:
            customers_by_id[c['id']] = c
        print(f"  Page {page}: {len(customers)} customers")
        if len(customers) < per_page:
            break
        page += 1
    
    print(f"  Customer list total: {len(customers_by_id)}")
    
    # Step 2: Also fetch customers from orders (catches guests + missing customers)
    print("\n[2/2] Fetching customers from orders (catching missing ones)...")
    page = 1
    orders_checked = 0
    customers_from_orders = 0
    
    while True:
        url = woo._url("/orders")
        response = woo.session.get(url, params={"page": page, "per_page": per_page}, timeout=60)
        if not response.ok:
            print(f"  Error fetching orders page {page}: {response.status_code}")
            break
        orders = response.json()
        if not orders:
            break
        
        for order in orders:
            orders_checked += 1
            customer_id = order.get('customer_id', 0)
            
            # Skip if already have this customer
            if customer_id in customers_by_id:
                continue
            
            # For registered customers not in list, fetch directly
            if customer_id > 0:
                cust_url = woo._url(f"/customers/{customer_id}")
                cust_response = woo.session.get(cust_url, timeout=30)
                if cust_response.ok:
                    cust = cust_response.json()
                    customers_by_id[customer_id] = cust
                    customers_from_orders += 1
            else:
                # Guest checkout - create pseudo-customer from order billing
                billing = order.get('billing', {})
                if billing.get('email'):
                    # Use negative order ID as pseudo-ID to avoid conflicts
                    pseudo_id = -order['id']
                    if pseudo_id not in customers_by_id:
                        customers_by_id[pseudo_id] = {
                            'id': f"GUEST-{order['id']}",
                            'email': billing.get('email'),
                            'first_name': billing.get('first_name'),
                            'last_name': billing.get('last_name'),
                            'username': '(guest)',
                            'role': 'guest',
                            'billing': billing,
                            'shipping': order.get('shipping', {}),
                            'date_created': order.get('date_created'),
                            'orders_count': 1,
                            'total_spent': order.get('total', '0.00'),
                        }
                        customers_from_orders += 1
        
        if len(orders) < per_page:
            break
        page += 1
    
    print(f"  Orders checked: {orders_checked}")
    print(f"  Additional customers found: {customers_from_orders}")
    
    all_customers = list(customers_by_id.values())
    print(f"\nTotal unique customers: {len(all_customers)}")
    return all_customers


def get_headers():
    """Return column headers."""
    return [
        "ID", "Email", "First_Name", "Last_Name", "Username", "Role",
        "Company", "Phone",
        "Billing_Address_1", "Billing_Address_2", "Billing_City", 
        "Billing_State", "Billing_Zip", "Billing_Country",
        "Shipping_Company", "Shipping_Phone",
        "Shipping_Address_1", "Shipping_Address_2", "Shipping_City",
        "Shipping_State", "Shipping_Zip", "Shipping_Country",
        "Date_Created", "Orders_Count", "Total_Spent"
    ]


def customer_to_row(c):
    """Convert a customer dict to a row of values."""
    billing = c.get('billing', {})
    shipping = c.get('shipping', {})
    
    return [
        c.get('id'),
        c.get('email'),
        c.get('first_name') or billing.get('first_name'),
        c.get('last_name') or billing.get('last_name'),
        c.get('username'),
        c.get('role'),
        billing.get('company'),
        billing.get('phone'),
        billing.get('address_1'),
        billing.get('address_2'),
        billing.get('city'),
        billing.get('state'),
        billing.get('postcode'),
        billing.get('country'),
        shipping.get('company'),
        shipping.get('phone'),
        shipping.get('address_1'),
        shipping.get('address_2'),
        shipping.get('city'),
        shipping.get('state'),
        shipping.get('postcode'),
        shipping.get('country'),
        c.get('date_created', '')[:10] if c.get('date_created') else '',
        c.get('orders_count', 0),
        c.get('total_spent', '0.00'),
    ]


def export_to_csv(customers, timestamp):
    """Export customers to CSV file."""
    headers = get_headers()
    filename = f"WooCommerce_Customers_{timestamp}.csv"
    filepath = os.path.join(os.path.dirname(__file__), filename)
    
    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for c in customers:
            writer.writerow(customer_to_row(c))
    
    print(f"\n[OK] CSV created: {filename}")
    return filepath


def export_to_excel(customers, timestamp):
    """Export customers to Excel file with formatting."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("Installing openpyxl...")
        os.system("pip install openpyxl")
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "WooCommerce Customers"
    
    headers = get_headers()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    ws.freeze_panes = "A2"
    
    # Write data
    for row, c in enumerate(customers, 2):
        for col, value in enumerate(customer_to_row(c), 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
    
    # Auto-adjust columns
    for col in range(1, len(headers) + 1):
        max_length = len(str(headers[col-1]))
        for row in range(2, min(len(customers) + 2, 50)):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 40)
    
    # Summary sheet
    ws_summary = wb.create_sheet("Summary")
    ws_summary["A1"] = "WooCommerce Customer Export"
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary["A3"] = "Export Date:"
    ws_summary["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws_summary["A4"] = "Total Customers:"
    ws_summary["B4"] = len(customers)
    ws_summary["A5"] = "Website:"
    ws_summary["B5"] = "woodyspaper.com"
    
    # Count by role
    roles = {}
    for c in customers:
        role = c.get('role', 'unknown')
        roles[role] = roles.get(role, 0) + 1
    
    ws_summary["A7"] = "By Role:"
    ws_summary["A7"].font = Font(bold=True)
    row = 8
    for role, count in sorted(roles.items(), key=lambda x: -x[1]):
        ws_summary[f"A{row}"] = f"  {role}:"
        ws_summary[f"B{row}"] = count
        row += 1
    
    filename = f"WooCommerce_Customers_{timestamp}.xlsx"
    filepath = os.path.join(os.path.dirname(__file__), filename)
    wb.save(filepath)
    
    print(f"[OK] Excel created: {filename}")
    return filepath


def main():
    # Parse command line argument
    format_arg = sys.argv[1].lower() if len(sys.argv) > 1 else 'csv'
    
    if format_arg not in ['csv', 'excel', 'both']:
        print("Usage: python export_woo_customers.py [csv|excel|both]")
        print("  csv   - Export to CSV (default)")
        print("  excel - Export to Excel with formatting")
        print("  both  - Export to both formats")
        return
    
    # Fetch customers
    customers = fetch_all_customers()
    
    if not customers:
        print("No customers found!")
        return
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Export based on format
    if format_arg == 'csv':
        export_to_csv(customers, timestamp)
    elif format_arg == 'excel':
        export_to_excel(customers, timestamp)
    elif format_arg == 'both':
        export_to_csv(customers, timestamp)
        export_to_excel(customers, timestamp)
    
    print(f"\nTotal customers exported: {len(customers)}")


if __name__ == "__main__":
    main()
